import os
from os import listdir
from os.path import isfile, join
import argparse
import yaml
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from cdisc_library_client import CDISCLibraryClient
import pandas as pd
import requests

def load_yaml_files(directory):

  files = [join(directory,f) for f in listdir(directory) if isfile(join(directory, f)) and os.path.splitext(f)[1] == '.yaml']
  print(f"\nGetting {len(files)} YAML files from {directory}")
  collection_list = []
  for f in files:
      with open(f, 'r') as stream:
          data_loaded = yaml.safe_load(stream)
      collection_list.append(data_loaded)
  return collection_list

def get_bc_list(client, cosmos_api_version,):
    cosmos_api_version = "v2"
    all_bc_list = client.get_bc_latest_biomedicalconcepts(cosmos_api_version)
    return all_bc_list

def get_bcs(client, cosmos_api_version, all_bc_list, count=1000):
    print(f"\nGetting {len(all_bc_list)} Biomedical Concepts from the CDISC Library")
    bc_list = []
    for bc in all_bc_list:
        href = bc['href']
        biomedicalConceptId = href.split("/")[-1]
        biomedicalConcept = client.get_bc_latest_biomedicalconcept(cosmos_api_version, biomedicalConceptId)
        bc_list.append(biomedicalConcept)
        if len(bc_list) == count:
            break
    return bc_list

def string_from_list(list):
    return ';'.join(str(item) for item in list)

def create_string(obj, key):
    object_key = obj.get(key, [])
    list = []
    for value in object_key:
        list.append(value)
    return string_from_list(list)

def create_parent_string(bc, parent, key):
    parent = bc.get(parent, [])
    list = []
    for p in parent:
        list.append(p.get(key))
    return string_from_list(list)

def get_bc_data(bc):
    if bc.get('_links'):
        package_date = bc.get("_links", "").get("parentPackage", "").get("href", "").split("/")[-2]
        if bc['_links'].get("parentBiomedicalConcept"):
            parent_bc_id = bc['_links'].get("parentBiomedicalConcept", "").get("href", "").split("/")[-1]
        else:
            parent_bc_id = ""
    else:
        package_date = bc.get("packageDate", "")
        parent_bc_id = bc.get("parentConceptId", "")
    short_name = bc.get("shortName", "")
    bc_id = bc.get("conceptId", "")
    ncit_code = bc.get("href", "").split("/")[-1]
    bc_categories = create_string(bc, "categories")
    synonyms = create_string(bc, "synonyms")
    result_scales = create_string(bc, "resultScales")
    definition = bc.get("definition", "")
    systems = create_parent_string(bc, "coding", "system")
    system_names = create_parent_string(bc, "coding", "systemName")
    codes = create_parent_string(bc, "coding", "code")
    return [package_date, short_name, bc_id, ncit_code, parent_bc_id, bc_categories, synonyms, result_scales, definition, systems, system_names, codes]

def get_bc_dec_data(bc):
    decs = bc.get("dataElementConcepts", [])
    dec_list = []
    for d in decs:
        dec_id = d.get("conceptId", "")
        ncit_dec_code = d.get("href", "").split("/")[-1]
        dec_label = d.get("shortName", "")
        data_type = d.get("dataType", "")
        example_set = create_string(d, "exampleSet")
        dec_list.append([dec_id, ncit_dec_code, dec_label, data_type, example_set])
    return dec_list

def get_bc_row(bc):
    bc_data = get_bc_data(bc)
    bc_dec_data = get_bc_dec_data(bc)
    return bc_data, bc_dec_data

def process_biomedical_concepts(bc_list, headers_bc_core, headers_bc_dec):
  headers_bc = headers_bc_core + headers_bc_dec
  empty_dec = ["", "", "", "", ""]
  bc_cat = set()
  bc_rows = []
  for bc in bc_list:
    bc_cat.update(bc.get("categories", []))
    bc_data, dec_data = get_bc_row(bc)
    if len(dec_data) == 0:
      bc_rows.append(bc_data + empty_dec)
    for dec in dec_data:
      bc = bc_data + dec
      bc_rows.append(bc)

  df = pd.DataFrame(bc_rows, columns=headers_bc)
  df = df.sort_values(by='short_name')
  df_cat = pd.DataFrame(list(bc_cat), columns=['bc_categories'])
  df_cat = df_cat.sort_values(by='bc_categories')
  return df, df_cat

def get_hierarchy_path(df, bc_id, visited=None):
    if visited is None:
        visited = set()

    if bc_id in visited or pd.isna(bc_id):
        return [], -1

    visited.add(bc_id)
    bc_row = df[df['bc_id'] == bc_id]
    if len(bc_row) == 0:
        return [], -1

    current_bc = bc_row['bc_short_name_id'].iloc[0]
    parent_id = bc_row['parent_bc_id'].iloc[0]

    if pd.isna(parent_id):
        return [current_bc], 0

    parent_path, parent_level = get_hierarchy_path(df, parent_id, visited)
    return (parent_path + [current_bc], parent_level + 1) if parent_path else ([current_bc], 0)

def create_bc_hierarchy(df, headers_bc_hierarchy):
    df_base = df[['short_name', 'bc_id', 'parent_bc_id',
                      'bc_categories', 'synonyms', 'result_scales', 'definition']].drop_duplicates()

    # Process data elements
    data_elements = df.groupby('bc_id')[['bc_id', 'dec_label', 'dec_id']].apply(
        lambda g: ';'.join(sorted([f"{label} ({id})"
                                 for label, id in zip(g['dec_label'], g['dec_id'])
                                 if pd.notna(label) and pd.notna(id)]))
    ).reset_index(name='bc_data_elements')


    # Add a column that counts the number of non-empty dec_id per bc_id, excluding rows where all dec_id values are empty
    df_new = df
    df_new['count_dec'] = 1 * (pd.notna(df['dec_id']) & (df['dec_id'] != ''))
    dec_id_counts = df_new.groupby('bc_id')['count_dec'].sum().reset_index(name='dec_n')

    # Merge the count data with data_elements
    data_elements = pd.merge(data_elements, dec_id_counts, on='bc_id', how='left')

    # Merge base data with data elements
    df_processed = pd.merge(df_base, data_elements, on='bc_id', how='left')

    df_processed = df_processed.rename(columns={
        'short_name': 'bc_short_name',
        'synonyms': 'bc_synonyms',
        'definition': 'bc_definition'
    })

    # Create bc_short_name_id
    df_processed['bc_short_name_id'] = df_processed.apply(
        lambda x: f"{x['bc_short_name']} ({x['bc_id']})", axis=1
    )

    # Process hierarchies
    hierarchies = {
        bc_id: get_hierarchy_path(df_processed, bc_id)
        for bc_id in df_processed['bc_id'] if not pd.isna(bc_id)
    }

    # Add hierarchy columns
    df_processed['bc_hierarchy_level'] = df_processed['bc_id'].map(
        lambda x: hierarchies.get(x, ([], 0))[1]
    )
    df_processed['bc_hierarchy_full'] = df_processed['bc_id'].map(
        lambda x: '; '.join(hierarchies.get(x, ([], 0))[0])
    )

    df_processed = df_processed[headers_bc_hierarchy]

    # Sort bc_short_name
    df_processed = df_processed.sort_values(['bc_short_name'])
    return df_processed

def write_biomedical_concepts_to_excel(workbook, sheetname, df):

  print(f"Writing {sheetname} sheet")
  NCI_LINK = "https://evsexplore.semantics.cancer.gov/evsexplore/concept/ncit/"
  headers_bc = [df.columns[x] for x in range(df.shape[1])]
  ws = workbook[sheetname]
  ft1 = Font(color='0000FF', underline='single')
  thin_border = Border(left=Side(style='thin', color='d3d3d3'),
                       right=Side(style='thin', color='d3d3d3'),
                       top=Side(style='thin', color='d3d3d3'),
                       bottom=Side(style='thin', color='d3d3d3'))
  for row_num, row_data in enumerate(df.values, 2):
    for col_num, col_data in enumerate(row_data, 1):
      if headers_bc[col_num-1] in ['ncit_code', 'parent_bc_id', 'ncit_dec_code']: # add hyperlink to ncit_code, parent_bc_id, ncit_dec_code
        try:
          if col_data.strip() != "" and col_data.strip()[:3] != "NEW":
            ws.cell(row=row_num, column=col_num).value = col_data.strip()
            ws.cell(row=row_num, column=col_num).font = ft1
            ws.cell(row=row_num, column=col_num).hyperlink = NCI_LINK + col_data.strip()
          elif col_data.strip()[:3] == "NEW":
            ws.cell(row=row_num, column=col_num).value = col_data
        except:
          ws.cell(row=row_num, column=col_num).value = col_data
      else:
        ws.cell(row=row_num, column=col_num).value = col_data
      if headers_bc[col_num-1] in ['short_name', 'bc_categories', 'definition', 'example_set']:
        ws.cell(row=row_num, column=col_num).alignment = Alignment(wrap_text=True)

      ws.cell(row=row_num, column=col_num).border = thin_border

  ws.auto_filter.ref = ws.dimensions
  return workbook

def write_dataframe_to_excel(workbook, sheetname, df):

  print(f"Writing {sheetname} sheet")
  headers_bc = [df.columns[x] for x in range(df.shape[1])]
  ws = workbook[sheetname]

  thin_border = Border(left=Side(style='thin', color='d3d3d3'),
                       right=Side(style='thin', color='d3d3d3'),
                       top=Side(style='thin', color='d3d3d3'),
                       bottom=Side(style='thin', color='d3d3d3'))

  for row_num, row_data in enumerate(df.values, 2):
    for col_num, col_data in enumerate(row_data, 1):
      ws.cell(row=row_num, column=col_num).value = col_data
      ws.cell(row=row_num, column=col_num).border = thin_border
      if headers_bc[col_num-1] in ['bc_short_name', 'bc_short_name_id', 'bc_categories', 'bc_definition', 'bc_hierarchy_full']:
        ws.cell(row=row_num, column=col_num).alignment = Alignment(wrap_text=True)

  return workbook

def update_readme(workbook, sheetname, source, date, count):

  print(f"Writing {sheetname} sheet")
  ws = workbook[sheetname]

  if source.lower() == "yaml":
      intro_text = f"This spreadsheet contains CDISC Biomedical Concepts loaded from {count} YAML files to be released on {date}.\nThe spreadsheet contains {count} unique CDISC Biomedical Concepts.\nThe image on the right shows the relation between Biomedical Concepts and SDTM Dataset Specializations.\nOnly a few attributes are shown in the image."
  else:
      intro_text = f"This spreadsheet contains the latest versions of CDISC Biomedical Concepts in the CDISC Library as of {date}.\nThere are currently {count} unique CDISC Biomedical Concepts in the CDISC Library.\nThe image  in the right shows the relation between Biomedical Concepts and SDTM Dataset Specializations. \nOnly a few attributes are shown in the image."

  ws.cell(row=1, column=1).value = intro_text
  ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)

  return workbook


def set_cmd_line_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("-s", "--source", required=True, default="API", help="Input source (YAML/API)", dest="source")
    parser.add_argument("-y", "--directory", help="Input folder with YAML files", dest="directory")
    parser.add_argument("-o", "--excel_file", default="cdisc_biomedical_concepts_latest.xlsx", help="Excel file to write the Biomedical Concepts to", dest="excel_file")
    parser.add_argument("-d", "--date, required=True", help="Latest package date", dest="bc_date")
    args = parser.parse_args()
    return args

def main():

    # Define the headers for the Biomedical Concepts
    HEADERS_BC_CORE = ["package_date", "short_name", "bc_id", "ncit_code", "parent_bc_id",
                       "bc_categories", "synonyms", "result_scales", "definition", "system", "system_name", "code",]
    HEADERS_BC_DEC = ["dec_id", "ncit_dec_code", "dec_label", "data_type", "example_set"]
    HEADERS_BC = HEADERS_BC_CORE + HEADERS_BC_DEC

    HEADERS_BC_HIERARCHY = ['bc_short_name','bc_id', 'bc_short_name_id', 'parent_bc_id',
                    'bc_categories', 'bc_synonyms', 'result_scales', 'bc_definition',
                    'bc_hierarchy_level', 'bc_hierarchy_full', 'dec_n']

    BC_TEMPLATE = "templates/cdisc_biomedical_concepts_latest_template.xlsx"

    args = set_cmd_line_args()

    if args.source.lower() == "api":
        api_key = os.environ.get("CDISC_LIBRARY_API_KEY")
        base_api_url = os.environ.get("CDISC_LIBRARY_API_URL")
        if not api_key or not base_api_url:
            print("Please set the CDISC_LIBRARY_API_KEY and CDISC_LIBRARY_API_URL environment variables.")
            return

        cosmos_api_version = "v2"
        client = CDISCLibraryClient(api_key=api_key, base_api_url=base_api_url)

        print(f"URL: {base_api_url}")
        if "dev" in base_api_url:
            client._session.verify = False
            requests.urllib3.disable_warnings()

    if args.source.lower() == "yaml":
        if args.directory is None:
            print("Please provide a directory with YAML files using -y or --directory")
            return
        bc_list = load_yaml_files(args.directory)
    else:
        all_bc_list = get_bc_list(client, cosmos_api_version)
        bc_list = get_bcs(client, cosmos_api_version, all_bc_list)

    script_path = os.path.dirname(os.path.realpath(__file__))
    bc_template = os.path.join(script_path, BC_TEMPLATE)

    df, df_cat = process_biomedical_concepts(bc_list, HEADERS_BC_CORE, HEADERS_BC_DEC)

    workbook = Workbook()
    workbook = load_workbook(bc_template)
    workbook = update_readme(workbook, "ReadMe", args.source, args.bc_date, len(bc_list))
    workbook = write_biomedical_concepts_to_excel(workbook, "Biomedical Concepts", df)
    workbook = write_dataframe_to_excel(workbook, "Categories", df_cat)
    df_hierarchy = create_bc_hierarchy(df, HEADERS_BC_HIERARCHY)
    workbook = write_dataframe_to_excel(workbook, "BC Hierarchy", df_hierarchy)
    workbook.save(args.excel_file)
    print(f"Excel file saved as {args.excel_file}")

    csv_file = args.excel_file.replace(".xlsx", ".csv")
    df.to_csv(csv_file, index=False, columns=HEADERS_BC)
    print(f"CSV file saved as {csv_file}")

    csv_file = args.excel_file.replace(".xlsx", "_hierarchy.csv")
    df_hierarchy.to_csv(csv_file, index=False)
    print(f"CSV file saved as {csv_file}")

if __name__ == "__main__":
    main()
