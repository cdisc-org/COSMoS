import argparse
import yaml
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from os import listdir
from os.path import isfile, join

def load_yaml_files(directory):

  files = [join(directory,f) for f in listdir(directory) if isfile(join(directory, f)) and os.path.splitext(f)[1] == '.yaml']
  print(f"\nGetting {len(files)} YAML files from {directory}")
  collection_list = []
  for f in files:
      with open(f, 'r') as stream:
          data_loaded = yaml.safe_load(stream)
      collection_list.append(data_loaded)
  return collection_list

def string_from_boolean(variable, boolean):
    if variable.get(boolean) == None:
      return ""
    else:
      if variable.get(boolean):
        return "Y"
      elif not variable.get(boolean):
        return "N"
      else:
        return ""

def string_from_list(list):
    return ';'.join(str(item) for item in list)

def create_string(obj, key):
    object_key = obj.get(key, [])
    list = []
    for value in object_key:
        list.append(value)
    return string_from_list(list)

def get_collection_data(collection):
    package_date = collection.get("packageDate", "")
    bc_id = collection.get("biomedicalConceptId", "")
    vlm_group_id = collection.get("sdtmDatasetSpecializationId", "")
    standard = collection.get("standard", "")
    standard_start_version = collection.get("standardStartVersion", "")
    standard_end_version = collection.get("standardEndVersion", "")
    domain = collection.get("domain", "")
    collection_group_id = collection.get("collectionSpecializationId", "")
    implementation_option = collection.get("implementationOption", "")
    scenario = collection.get("scenario", "")
    short_name = collection.get("shortName", "")
    return [package_date, bc_id, vlm_group_id, standard, standard_start_version, standard_end_version, domain,
            collection_group_id, implementation_option, scenario, short_name]

def get_collection_item_data(collection):
    items = collection.get("items", [])
    item_list = []
    for v in items:
        collection_item = v.get("name", "")
        variable_name = v.get("variableName", "")
        dec_id = v.get("dataElementConceptId", "")
        question_text = v.get("questionText", "")
        prompt = v.get("prompt", "")
        order_number = v.get("orderNumber", None)
        mandatory_variable = string_from_boolean(v, "mandatoryVariable")
        data_type = v.get("dataType", "")
        length = v.get("length", None)
        significant_digits = v.get("significantDigits", None)
        display_hidden = string_from_boolean(v, "displayHidden")
        if v.get("codelist"):
          codelist = v.get("codelist").get("conceptId", "")
          codelist_submission_value = v.get("codelist").get("submissionValue", "")
        else:
          codelist = ""
          codelist_submission_value = ""

        valueListDisplay = []
        valueListValue = []
        if v.get("valueList"):
          for vl in v.get("valueList"):
            valueListDisplay.append(vl.get("displayValue", ""))
            valueListValue.append(vl.get("value", ""))
        else:
          valueListDisplay = []
          valueListValue = []
        value_display_list = string_from_list(valueListDisplay)
        value_list = string_from_list(valueListValue)

        selection_type = v.get("selectionType", "")
        if v.get("prepopulatedValue"):
            prepopulated_term = v.get("prepopulatedValue").get("value", "")
            prepopulated_code = v.get("prepopulatedValue").get("conceptId", "")
        else:
            prepopulated_term = ""
            prepopulated_code = ""

        if v.get("sdtmTarget"):
          sdtm_annotation = v.get("sdtmTarget").get("sdtmAnnotation", "")
          sdtmVariables = v.get("sdtmTarget").get("sdtmVariables", [])
          if sdtmVariables:
              sdtm_target_variable = create_string(v.get("sdtmTarget"), "sdtmVariables")
          else:
              sdtm_target_variable = ""
          sdtm_mapping = v.get("sdtmTarget").get("sdtmTargetMapping", "")
        else:
          sdtm_annotation = ""
          sdtm_target_variable = ""
          sdtm_mapping = ""

        item_list.append([collection_item, variable_name, dec_id, question_text, prompt, order_number, mandatory_variable,
                          data_type, length, significant_digits, display_hidden, codelist, codelist_submission_value,
                          value_list, value_display_list, selection_type, prepopulated_term, prepopulated_code,
                          sdtm_target_variable, sdtm_annotation, sdtm_mapping])
    return item_list

def get_collection_row(collection):
    collection_data = get_collection_data(collection)
    collection_item_data = get_collection_item_data(collection)
    return collection_data, collection_item_data

def process_collection_dataset_specializations(collection_list, headers_collection_core, headers_collection_variable):
  headers_collection = headers_collection_core + headers_collection_variable
  collection_domain = set()
  collection_rows = []
  for collection in collection_list:
    collection_domain.add(collection.get('domain'))
    collection_data, item_data = get_collection_row(collection)
    for item in item_data:
      collection = collection_data + item
      collection_rows.append(collection)

  df = pd.DataFrame(collection_rows, columns=headers_collection)
  df = df.sort_values(by=['domain', 'collection_group_id', 'order_number'])
  df_domain = pd.DataFrame(list(collection_domain), columns=['domain'])
  df_domain = df_domain.sort_values(by='domain')
  return df, df_domain

def write_collection_dataset_specializations_to_excel(workbook, sheetname, df):

  print(f"Writing {sheetname} sheet")
  NCI_LINK = "https://evsexplore.semantics.cancer.gov/evsexplore/concept/ncit/"
  headers_bc = [df.columns[x] for x in range(df.shape[1])]
  ws = workbook[sheetname]
  ft1 = Font(color='0000FF', underline='single')
  thin_border = Border(left=Side(style='thin', color='d3d3d3'),
                       right=Side(style='thin', color='d3d3d3'),
                       top=Side(style='thin', color='d3d3d3'),
                       bottom=Side(style='thin', color='d3d3d3'))

  for row_num, row_data in enumerate(df.values, 2):
    for col_num, col_data in enumerate(row_data, 1):
      if headers_bc[col_num-1] in ['bc_id', 'dec_id', 'codelist', 'prepopulated_code']: # add hyperlink
        try:
          if col_data.strip() != "" and col_data.strip()[:3] != "NEW":
            ws.cell(row=row_num, column=col_num).value = col_data.strip()
            ws.cell(row=row_num, column=col_num).font = ft1
            ws.cell(row=row_num, column=col_num).hyperlink = NCI_LINK + col_data.strip()
          elif col_data.strip()[:3] == "NEW":
            ws.cell(row=row_num, column=col_num).value = col_data
        except:
          ws.cell(row=row_num, column=col_num).value = col_data
      else:
        ws.cell(row=row_num, column=col_num).value = col_data
      if headers_bc[col_num-1] in ['short_name', 'question_text', 'prompt', 'value_list', 'value_display_list', 'sdtm_annotation', 'sdtm_mapping']:
        ws.cell(row=row_num, column=col_num).alignment = Alignment(wrap_text=True)

      ws.cell(row=row_num, column=col_num).border = thin_border

  ws.auto_filter.ref = ws.dimensions
  return workbook

def write_dataframe_to_excel(workbook, sheetname, df):

  print(f"Writing {sheetname} sheet")
  ws = workbook[sheetname]

  thin_border = Border(left=Side(style='thin', color='d3d3d3'),
                       right=Side(style='thin', color='d3d3d3'),
                       top=Side(style='thin', color='d3d3d3'),
                       bottom=Side(style='thin', color='d3d3d3'))

  for row_num, row_data in enumerate(df.values, 2):
    for col_num, col_data in enumerate(row_data, 1):
      ws.cell(row=row_num, column=col_num).value = col_data
      ws.cell(row=row_num, column=col_num).border = thin_border

  return workbook

def set_cmd_line_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("-s", "--source", required=True, default="API", help="Input source (YAML/API)", dest="source")
    parser.add_argument("-y", "--directory", required=True, help="Input folder with YAML files", dest="directory")
    parser.add_argument("-o", "--excel_file", default="cdisc_collection_dataset_specializations_latest.xlsx", help="Excel file to write the Collection Dataset Specializations to", dest="excel_file")
    parser.add_argument("-d", "--date, required=True", help="Latest package date", dest="bc_date")
    args = parser.parse_args()
    return args

def main():

    # Define the headers for the Collection Dataset Specializations
    HEADERS_COLLECTION_CORE = ["package_date", "bc_id", "vlm_group_id", "standard", "standard_start_version", "standard_end_version", "domain",
                               "collection_group_id", "implementation_option", "scenario", "short_name"]
    HEADERS_COLLECTION_ITEM = ["collection_item", "variable_name", "dec_id", "question_text", "prompt", "order_number", "mandatory_variable",
                               "data_type", "length", "significant_digits", "display_hidden", "codelist", "codelist_submission_value",
                               "value_list", "value_display_list", "selection_type", "prepopulated_term", "prepopulated_code",
                               "sdtm_target_variable", "sdtm_annotation", "sdtm_mapping"]
    HEADERS_COLLECTION = HEADERS_COLLECTION_CORE + HEADERS_COLLECTION_ITEM

    TEMPLATE_COLLECTION = "templates/cdisc_collection_dataset_specializations_latest_template.xlsx"

    args = set_cmd_line_args()

    if args.source.lower() == "yaml":
        if args.directory is None:
            print("Please provide a directory with YAML files using -y or --directory")
            return
        collection_list = load_yaml_files(args.directory)
    else:
        print("Only YAML source is supported at this time.")
        return

    script_path = os.path.dirname(os.path.realpath(__file__))
    collection_template = os.path.join(script_path, TEMPLATE_COLLECTION)

    df, df_domain = process_collection_dataset_specializations(collection_list, HEADERS_COLLECTION_CORE, HEADERS_COLLECTION_ITEM)

    workbook = Workbook()
    workbook = load_workbook(collection_template)
    # workbook = update_readme(workbook, "ReadMe", args.bc_date, len(collection_list))
    workbook = write_collection_dataset_specializations_to_excel(workbook, "Collection Specializations", df)
    workbook = write_dataframe_to_excel(workbook, "Domains", df_domain)
    workbook.save(args.excel_file)
    print(f"Excel file saved as {args.excel_file}")

    csv_file = args.excel_file.replace(".xlsx", ".csv")
    df.to_csv(csv_file, index=False, columns=HEADERS_COLLECTION)
    print(f"CSV file saved as {csv_file}")

if __name__ == "__main__":
    main()

# This script creates an Excel file with Collection Dataset Specializations from YAML files.
# Usage: python create_cosmos_collection_excel.py -s <source> -y <directory> -o <excel_file> -d <date>
# Source can be 'YAML'. Provide a directory with YAML files.
# If 'API', the script will use the CDISC Library API to get the latest Collection Dataset Specializations.
# The output will be saved in an Excel file and a CSV file.
# The Excel file will contain sheets for Collection Specializations and Domains.
