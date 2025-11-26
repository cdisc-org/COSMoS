import argparse
import yaml
from datetime import date
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from os import listdir
from os.path import isfile, join

"""
This script creates an Excel file with Collection Dataset Specializations from YAML files.
Usage: python create_cosmos_crf_excel.py -s <source> -y <directory> -o <excel_file> -d <date>
Source can be 'YAML'. Provide a directory with YAML files.
If 'API', the script will use the CDISC Library API to get the latest Collection Dataset Specializations.
The output will be saved in an Excel file and a CSV file.
The Excel file will contain sheets for Collection Specializations and Domains.
"""


def load_yaml_files(directory):

    files = [
        join(directory, f)
        for f in listdir(directory)
        if isfile(join(directory, f)) and os.path.splitext(f)[1] == '.yaml'
    ]
    print(f"\nGetting {len(files)} YAML files from {directory}")
    crf_list = []
    for f in files:
        with open(f, 'r') as stream:
            data_loaded = yaml.safe_load(stream)
        crf_list.append(data_loaded)
    return crf_list


def string_from_boolean(variable, boolean):
    if variable.get(boolean) is None:
        return ""
    else:
        if variable.get(boolean):
            return "Y"
        elif not variable.get(boolean):
            return "N"
        else:
            return ""


def string_from_list(list):
    return ';'.join(str(item) for item in list)


def create_string(obj, key):
    object_key = obj.get(key, [])
    list = []
    for value in object_key:
        list.append(value)
    return string_from_list(list)


def get_crf_data(crf):
    package_date = crf.get("packageDate", "")
    bc_id = crf.get("biomedicalConceptId", "")
    vlm_group_id = crf.get("sdtmDatasetSpecializationId", "")
    standard = crf.get("standard", "")
    standard_start_version = crf.get("standardStartVersion", "")
    standard_end_version = crf.get("standardEndVersion", "")
    domain = crf.get("domain", "")
    crf_group_id = crf.get("crfSpecializationId", "")
    implementation_option = crf.get("implementationOption", "")
    scenario = crf.get("scenario", "")
    categories = create_string(crf, "categories")
    short_name = crf.get("shortName", "")
    return [package_date, bc_id, vlm_group_id, standard, standard_start_version, standard_end_version, domain,
            crf_group_id, implementation_option, scenario, categories, short_name]


def get_crf_item_data(crf):
    items = crf.get("items", [])
    item_list = []
    for v in items:
        crf_item = v.get("name", "")
        variable_name = v.get("variableName", "")
        dec_id = v.get("dataElementConceptId", "")
        question_text = v.get("questionText", "")
        prompt = v.get("prompt", "")
        completion_instructions = v.get("completion_instructions", "")
        order_number = v.get("orderNumber", None)
        mandatory_variable = string_from_boolean(v, "mandatoryVariable")
        data_type = v.get("dataType", "")
        length = v.get("length", None)
        significant_digits = v.get("significantDigits", None)
        display_hidden = string_from_boolean(v, "displayHidden")
        derived_variable = string_from_boolean(v, "derivedVariable")
        derivation_description = v.get("derivationDescription", "")
        if v.get("codelist"):
            codelist = v.get("codelist").get("conceptId", "")
            codelist_submission_value = v.get("codelist").get("submissionValue", "")
        else:
            codelist = ""
            codelist_submission_value = ""

        valueListDisplay = []
        valueListValue = []
        if v.get("valueList"):
            for vl in v.get("valueList"):
                valueListDisplay.append(vl.get("displayValue", ""))
                valueListValue.append(vl.get("value", ""))
        else:
            valueListDisplay = []
            valueListValue = []
        value_display_list = string_from_list(valueListDisplay)
        value_list = string_from_list(valueListValue)

        selection_type = v.get("selectionType", "")
        if v.get("prepopulatedValue"):
            prepopulated_term = v.get("prepopulatedValue").get("value", "")
            prepopulated_code = v.get("prepopulatedValue").get("conceptId", "")
        else:
            prepopulated_term = ""
            prepopulated_code = ""

        if v.get("sdtmTarget"):
            sdtm_annotation = v.get("sdtmTarget").get("sdtmAnnotation", "")
            sdtmVariables = v.get("sdtmTarget").get("sdtmVariables", [])
            if sdtmVariables:
                sdtm_target_variable = create_string(v.get("sdtmTarget"), "sdtmVariables")
            else:
                sdtm_target_variable = ""
            sdtm_mapping = v.get("sdtmTarget").get("sdtmTargetMapping", "")
        else:
            sdtm_annotation = ""
            sdtm_target_variable = ""
            sdtm_mapping = ""

        item_list.append([
            crf_item, variable_name, dec_id, question_text, prompt, completion_instructions, order_number,
            mandatory_variable, data_type, length, significant_digits, display_hidden, derived_variable,
            derivation_description, codelist, codelist_submission_value, value_list, value_display_list, selection_type,
            prepopulated_term, prepopulated_code, sdtm_target_variable, sdtm_annotation, sdtm_mapping
        ])
    return item_list


def get_crf_row(crf):
    crf_data = get_crf_data(crf)
    crf_item_data = get_crf_item_data(crf)
    return crf_data, crf_item_data


def process_crf_dataset_specializations(crf_list, headers_crf_core, headers_crf_variable):
    headers_crf = headers_crf_core + headers_crf_variable
    crf_domain = set()
    crf_rows = []
    for crf in crf_list:
        crf_domain.add(crf.get('domain'))
        crf_data, item_data = get_crf_row(crf)
        for item in item_data:
            crf = crf_data + item
            crf_rows.append(crf)

    df = pd.DataFrame(crf_rows, columns=headers_crf)
    df = df.sort_values(by=['domain', 'crf_group_id', 'order_number'])
    df_domain = pd.DataFrame(list(crf_domain), columns=['domain'])
    df_domain = df_domain.sort_values(by='domain')
    return df, df_domain


def update_readme(workbook, sheetname, source, date, count):

    print(f"Writing {sheetname} sheet")
    ws = workbook[sheetname]
    if source.lower() == "yaml":
        intro_text = (
            (
                f"This spreadsheet contains CDISC CRF Specializations loaded from "
                f"{count} YAML files as of {date}.\n"
                f"The spreadsheet contains {count} unique CDISC CRF Specializations.\n"
                "The image on the right shows the relation between Biomedical Concepts, "
                "CRF Specializations and SDTM Dataset Specializations. "
                "Only a few attributes are shown in the image."
            )
        )
    else:
        intro_text = (
            f"This spreadsheet contains the latest versions of CDISC CRF Specializations "
            f"in the CDISC Library as of {date}.\n"
            f"There are currently {count} unique CDISC CRF Specializations in the CDISC Library.\n"
            "The image on the right shows the relation between Biomedical Concepts, "
            "CRF Specializations and SDTM Dataset Specializations. "
            "Only a few attributes are shown in the image."
        )

    ws.cell(row=1, column=1).value = intro_text
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)

    return workbook


def write_crf_dataset_specializations_to_excel(workbook, sheetname, df):

    print(f"Writing {sheetname} sheet")
    NCI_LINK = "https://evsexplore.semantics.cancer.gov/evsexplore/concept/ncit/"
    headers_bc = [df.columns[x] for x in range(df.shape[1])]
    ws = workbook[sheetname]
    ft1 = Font(color='0000FF', underline='single')
    thin_border = Border(
        left=Side(style='thin', color='d3d3d3'),
        right=Side(style='thin', color='d3d3d3'),
        top=Side(style='thin', color='d3d3d3'),
        bottom=Side(style='thin', color='d3d3d3')
    )

    for row_num, row_data in enumerate(df.values, 2):
        for col_num, col_data in enumerate(row_data, 1):
            if headers_bc[col_num - 1] in ['bc_id', 'dec_id', 'codelist', 'prepopulated_code']:  # add hyperlink
                try:
                    if col_data.strip() != "" and col_data.strip()[:3] != "NEW":
                        ws.cell(row=row_num, column=col_num).value = col_data.strip()
                        ws.cell(row=row_num, column=col_num).font = ft1
                        ws.cell(row=row_num, column=col_num).hyperlink = NCI_LINK + col_data.strip()
                    elif col_data.strip()[:3] == "NEW":
                        ws.cell(row=row_num, column=col_num).value = col_data
                except Exception:
                    ws.cell(row=row_num, column=col_num).value = col_data
            else:
                ws.cell(row=row_num, column=col_num).value = col_data
            if headers_bc[col_num - 1] in [
                'short_name', 'question_text', 'prompt', 'complertion_instructions', 'derivation_description',
                'value_list', 'value_display_list', 'sdtm_annotation', 'sdtm_mapping'
            ]:
                ws.cell(row=row_num, column=col_num).alignment = Alignment(wrap_text=True)

            ws.cell(row=row_num, column=col_num).border = thin_border

    ws.auto_filter.ref = ws.dimensions
    return workbook


def write_dataframe_to_excel(workbook, sheetname, df):

    print(f"Writing {sheetname} sheet")
    ws = workbook[sheetname]

    thin_border = Border(
        left=Side(style='thin', color='d3d3d3'),
        right=Side(style='thin', color='d3d3d3'),
        top=Side(style='thin', color='d3d3d3'),
        bottom=Side(style='thin', color='d3d3d3')
    )

    for row_num, row_data in enumerate(df.values, 2):
        for col_num, col_data in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = col_data
            ws.cell(row=row_num, column=col_num).border = thin_border

    return workbook


def set_cmd_line_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("-s", "--source", required=True, default="API", help="Input source (YAML/API)", dest="source")
    parser.add_argument("-y", "--directory", required=True, help="Input folder with YAML files", dest="directory")
    parser.add_argument(
        "-o",
        "--excel_file",
        default="cdisc_crf_specializations_latest.xlsx",
        help="Excel file to write the CRF Specializations to",
        dest="excel_file"
    )
    parser.add_argument("-d", "--date, required=True", help="Latest package date", dest="bc_date")
    args = parser.parse_args()
    return args


def main():

    # Define the headers for the Collection Dataset Specializations
    HEADERS_CRF_CORE = [
        "package_date", "bc_id", "vlm_group_id", "standard", "standard_start_version",
        "standard_end_version", "domain", "crf_group_id", "implementation_option",
        "scenario", "categories", "short_name"
    ]
    HEADERS_CRF_ITEM = [
        "crf_item", "variable_name", "dec_id", "question_text", "prompt", "completion_instructions",
        "order_number", "mandatory_variable", "data_type", "length", "significant_digits",
        "display_hidden", 'derived_variable', 'derivation_description',
        "codelist", "codelist_submission_value", "value_list", "value_display_list",
        "selection_type", "prepopulated_term", "prepopulated_code", "sdtm_target_variable",
        "sdtm_annotation", "sdtm_mapping"
    ]
    HEADERS_CRF = HEADERS_CRF_CORE + HEADERS_CRF_ITEM

    TEMPLATE_CRF = "templates/cdisc_crf_specializations_latest_template.xlsx"

    args = set_cmd_line_args()

    if args.source.lower() == "yaml":
        if args.directory is None:
            print("Please provide a directory with YAML files using -y or --directory")
            return
        crf_list = load_yaml_files(args.directory)
    else:
        print(
            "Only YAML source is supported at this time."
        )
        return

    script_path = os.path.dirname(os.path.realpath(__file__))
    crf_template = os.path.join(script_path, TEMPLATE_CRF)

    df, df_domain = process_crf_dataset_specializations(
        crf_list,
        HEADERS_CRF_CORE,
        HEADERS_CRF_ITEM
    )
    current_date = date.today()

    workbook = Workbook()
    workbook = load_workbook(crf_template)
    workbook = update_readme(workbook, "ReadMe", args.source, current_date.isoformat(), len(crf_list))
    workbook = write_crf_dataset_specializations_to_excel(workbook, "CRF Specializations", df)
    workbook = write_dataframe_to_excel(workbook, "Domains", df_domain)
    workbook.save(args.excel_file)
    print(f"Excel file saved as {args.excel_file}")

    csv_file = args.excel_file.replace(".xlsx", ".csv")
    df.to_csv(csv_file, index=False, columns=HEADERS_CRF)
    print(f"CSV file saved as {csv_file}")


if __name__ == "__main__":
    main()
