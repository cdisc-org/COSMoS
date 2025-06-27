import os
from os import listdir
from os.path import isfile, join
import argparse
import yaml
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from cdisc_library_client import CDISCLibraryClient
import pandas as pd
import requests

"""
This script creates an Excel file with SDTM Dataset Specializations from YAML files or the CDISC Library API.
Usage: python create_cosmos_sdtm_excel.py -s <source> -y <directory> -o <excel_file> -d <date>
Source can be 'API' or 'YAML'. If 'YAML', provide a directory with YAML files.
If 'API', the script will use the CDISC Library API to get the latest SDTM Dataset Specializations.
The output will be saved in an Excel file and a CSV file.
The Excel file will contain sheets for SDTM Dataset Specializations and Domains.
The CSV file will contain the SDTM Dataset Specializations data.
The script also updates a ReadMe sheet with information about the source and date of the data.
"""


def load_yaml_files(directory):

  files = [join(directory,f) for f in listdir(directory) if isfile(join(directory, f)) and os.path.splitext(f)[1] == '.yaml']
  print(f"\nGetting {len(files)} YAML files from {directory}")
  collection_list = []
  for f in files:
      with open(f, 'r') as stream:
          data_loaded = yaml.safe_load(stream)
      collection_list.append(data_loaded)
  return collection_list

def get_sdtm_list(client, cosmos_api_version,):
    cosmos_api_version = "v2"
    all_sdtm_list = client.get_sdtm_latest_sdtm_datasetspecializations(cosmos_api_version)
    return all_sdtm_list

def get_sdtms(client, cosmos_api_version, all_sdtm_list, count=1000):
    print(f"\nGetting {len(all_sdtm_list)} SDTM Dataset Specializations from the CDISC Library")
    sdtm_list = []
    for sdtm in all_sdtm_list:
        href = sdtm['href']
        datasetSpecializationId = href.split("/")[-1]
        datasetSpecialization = client.get_sdtm_latest_sdtm_datasetspecialization(cosmos_api_version, datasetSpecializationId)
        sdtm_list.append(datasetSpecialization)
        if len(sdtm_list) == count:
            break
    return sdtm_list

def string_from_boolean(variable, boolean):
    if variable.get(boolean) == None:
      return ""
    else:
      if variable.get(boolean):
        return "Y"
      elif not variable.get(boolean):
        return "N"
      else:
        return ""

def string_from_list(list):
    return ';'.join(str(item) for item in list)

def create_string(obj, key):
    object_key = obj.get(key, [])
    list = []
    for value in object_key:
        list.append(value)
    return string_from_list(list)

def get_sdtm_data(sdtm):
    if sdtm.get('_links'):
        package_date = sdtm.get("_links", "").get("parentPackage", "").get("href", "").split("/")[-2]
        if sdtm['_links'].get("parentBiomedicalConcept"):
            bc_id = sdtm['_links'].get("parentBiomedicalConcept", "").get("href", "").split("/")[-1]
        else:
            bc_id = ""
    else:
        bc_id = sdtm.get("biomedicalConceptId", "")
        package_date = sdtm.get("packageDate", "")
    sdtmig_start_version = sdtm.get("sdtmigStartVersion", "")
    sdtmig_end_version = sdtm.get("sdtmigEndVersion", "")
    domain = sdtm.get("domain", "")
    vlm_source = sdtm.get("source", "")
    vlm_group_id = sdtm.get("datasetSpecializationId", "")
    short_name = sdtm.get("shortName", "")
    return [package_date, bc_id, sdtmig_start_version, sdtmig_end_version, domain, vlm_source, vlm_group_id, short_name]

def get_sdtm_variable_data(sdtm):
    variables = sdtm.get("variables", [])
    variable_list = []
    v_order = 0
    for v in variables:
        v_order = v_order + 1
        sdtm_variable = v.get("name", "")
        dec_id = v.get("dataElementConceptId", "")
        nsv_flag = string_from_boolean(v, "isNonStandard")
        if v.get("codelist"):
          codelist = v.get("codelist").get("conceptId", "")
          codelist_submission_value = v.get("codelist").get("submissionValue", "")
        else:
          codelist = ""
          codelist_submission_value = ""
        subset_codelist= v.get("subsetCodelist", "")
        value_list = create_string(v, "valueList")
        if v.get("assignedTerm"):
          assigned_term = v.get("assignedTerm").get("conceptId", "")
          assigned_value = v.get("assignedTerm").get("value", "")
        else:
          assigned_term = ""
          assigned_value = ""
        role = v.get("role", "")
        if v.get("relationship"):
          subject = v.get("relationship").get("subject", "")
          linking_phrase = v.get("relationship").get("linkingPhrase", "")
          predicate_term = v.get("relationship").get("predicateTerm", "")
          object = v.get("relationship").get("object", "")
        else:
          subject = ""
          linking_phrase = ""
          predicate_term = ""
          object = ""
        data_type = v.get("dataType", "")
        length = v.get("length", None)
        format = str(v.get("format", ""))
        significant_digits = v.get("significantDigits", None)
        mandatory_variable = string_from_boolean(v, "mandatoryVariable")
        mandatory_value = string_from_boolean(v, "mandatoryValue")
        origin_type = v.get("originType", "")
        origin_source = v.get("originSource", "")
        comparator = v.get("comparator", "")
        vlm_target = string_from_boolean(v, "vlmTarget")
        variable_list.append([v_order, sdtm_variable, dec_id, nsv_flag, codelist, codelist_submission_value, subset_codelist, value_list, assigned_term, assigned_value, role, subject, linking_phrase, predicate_term, object, data_type, length, format, significant_digits, mandatory_variable, mandatory_value, origin_type, origin_source, comparator, vlm_target])
    return variable_list

def get_sdtm_row(sdtm):
    sdtm_data = get_sdtm_data(sdtm)
    sdtm_variable_data = get_sdtm_variable_data(sdtm)
    return sdtm_data, sdtm_variable_data

def process_sdtm_dataset_specializations(sdtm_list, headers_sdtm_core, headers_sdtm_variable):
  headers_sdtm = headers_sdtm_core + ['v_order'] + headers_sdtm_variable
  sdtm_domain = set()
  sdtm_rows = []
  for sdtm in sdtm_list:
    sdtm_domain.add(sdtm.get('domain'))
    sdtm_data, variable_data = get_sdtm_row(sdtm)
    for variable in variable_data:
      sdtm = sdtm_data + variable
      sdtm_rows.append(sdtm)

  df = pd.DataFrame(sdtm_rows, columns=headers_sdtm)
  df = df.sort_values(by=['domain', 'vlm_group_id', 'v_order'])
  df = df.drop(columns=['v_order'])
  df_domain = pd.DataFrame(list(sdtm_domain), columns=['domain'])
  df_domain = df_domain.sort_values(by='domain')
  return df, df_domain

def update_readme(workbook, sheetname, source, date, count):

  print(f"Writing {sheetname} sheet")
  ws = workbook[sheetname]
  if source.lower() == "yaml":
      intro_text = f"This spreadsheet contains CDISC SDTM Dataset Specializations loaded from {count} YAML files to be released on {date}.\nThe spreadsheet contains {count} unique CDISC SDTM Dataset Specializations.\nThe image on the right shows the relation between Biomedical Concepts and SDTM Dataset Specializations.\nOnly a few attributes are shown in the image."
  else:
      intro_text = f"This spreadsheet contains the latest versions of CDISC SDTM Dataset Specializations in the CDISC Library as of {date}.\nThere are currently {count} unique CDISC SDTM Dataset Specializations in the CDISC Library.\nThe image on the right shows the relation between Biomedical Concepts and SDTM Dataset Specializations.\nOnly a few attributes are shown in the image."


  ws.cell(row=1, column=1).value = intro_text
  ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)

  return workbook

def write_sdtm_dataset_specializations_to_excel(workbook, sheetname, df):

  print(f"Writing {sheetname} sheet")
  NCI_LINK = "https://evsexplore.semantics.cancer.gov/evsexplore/concept/ncit/"
  headers_bc = [df.columns[x] for x in range(df.shape[1])]
  ws = workbook[sheetname]
  ft1 = Font(color='0000FF', underline='single')
  thin_border = Border(left=Side(style='thin', color='d3d3d3'),
                       right=Side(style='thin', color='d3d3d3'),
                       top=Side(style='thin', color='d3d3d3'),
                       bottom=Side(style='thin', color='d3d3d3'))

  for row_num, row_data in enumerate(df.values, 2):
    for col_num, col_data in enumerate(row_data, 1):
      if headers_bc[col_num-1] in ['bc_id', 'dec_id', 'codelist']: # add hyperlink to bc_id, dec_id, codelist
        try:
          if col_data.strip() != "" and col_data.strip()[:3] != "NEW":
            ws.cell(row=row_num, column=col_num).value = col_data.strip()
            ws.cell(row=row_num, column=col_num).font = ft1
            ws.cell(row=row_num, column=col_num).hyperlink = NCI_LINK + col_data.strip()
          elif col_data.strip()[:3] == "NEW":
            ws.cell(row=row_num, column=col_num).value = col_data
        except:
          ws.cell(row=row_num, column=col_num).value = col_data
      else:
        ws.cell(row=row_num, column=col_num).value = col_data
      if headers_bc[col_num-1] in ['short_name', 'linking_phrase']:
        ws.cell(row=row_num, column=col_num).alignment = Alignment(wrap_text=True)

      ws.cell(row=row_num, column=col_num).border = thin_border

  ws.auto_filter.ref = ws.dimensions
  return workbook

def write_dataframe_to_excel(workbook, sheetname, df):

  print(f"Writing {sheetname} sheet")
  ws = workbook[sheetname]

  thin_border = Border(left=Side(style='thin', color='d3d3d3'),
                       right=Side(style='thin', color='d3d3d3'),
                       top=Side(style='thin', color='d3d3d3'),
                       bottom=Side(style='thin', color='d3d3d3'))

  for row_num, row_data in enumerate(df.values, 2):
    for col_num, col_data in enumerate(row_data, 1):
      ws.cell(row=row_num, column=col_num).value = col_data
      ws.cell(row=row_num, column=col_num).border = thin_border

  return workbook

def set_cmd_line_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("-s", "--source", required=True, default="API", help="Input source (YAML/API)", dest="source")
    parser.add_argument("-y", "--directory", help="Input folder with YAML files", dest="directory")
    parser.add_argument("-o", "--excel_file", default="cdisc_sdtm_dataset_specializations_latest.xlsx", help="Excel file to write the SDTM Dataset Specializations to", dest="excel_file")
    parser.add_argument("-d", "--date, required=True", help="Latest package date", dest="bc_date")
    args = parser.parse_args()
    return args

def main():

    # Define the headers for the SDTM Dataset Specializations
    HEADERS_SDTM_CORE = ["package_date", "bc_id", "sdtmig_start_version", "sdtmig_end_version", "domain", "vlm_source", "vlm_group_id", "short_name"]
    HEADERS_SDTM_VARIABLE = ["sdtm_variable", "dec_id", "nsv_flag", "codelist", "codelist_submission_value", "subset_codelist",
                      "value_list", "assigned_term", "assigned_value", "role", "subject",
                      "linking_phrase", "predicate_term", "object", "data_type", "length", "format", "significant_digits",
                      "mandatory_variable", "mandatory_value", "origin_type", "origin_source", "comparator", "vlm_target"]
    HEADERS_SDTM = HEADERS_SDTM_CORE + HEADERS_SDTM_VARIABLE

    SDTM_TEMPLATE = "templates/cdisc_sdtm_dataset_specializations_latest_template.xlsx"

    args = set_cmd_line_args()

    if args.source.lower() == "api":
        api_key = os.environ.get("CDISC_LIBRARY_API_KEY")
        base_api_url = os.environ.get("CDISC_LIBRARY_API_URL")
        if not api_key or not base_api_url:
            print("Please set the CDISC_LIBRARY_API_KEY and CDISC_LIBRARY_API_URL environment variables.")
            return

        cosmos_api_version = "v2"
        client = CDISCLibraryClient(api_key=api_key, base_api_url=base_api_url)

        print(f"URL: {base_api_url}")
        if "dev" in base_api_url:
          client._session.verify = False
          requests.urllib3.disable_warnings()

    if args.source.lower() == "yaml":
        if args.directory is None:
            print("Please provide a directory with YAML files using -y or --directory")
            return
        sdtm_list = load_yaml_files(args.directory)
    else:
        all_sdtm_list = get_sdtm_list(client, cosmos_api_version)
        sdtm_list = get_sdtms(client, cosmos_api_version, all_sdtm_list)

    script_path = os.path.dirname(os.path.realpath(__file__))
    sdtm_template = os.path.join(script_path, SDTM_TEMPLATE)

    df, df_domain = process_sdtm_dataset_specializations(sdtm_list, HEADERS_SDTM_CORE, HEADERS_SDTM_VARIABLE)

    workbook = Workbook()
    workbook = load_workbook(sdtm_template)
    workbook = update_readme(workbook, "ReadMe", args.source, args.bc_date, len(sdtm_list))
    workbook = write_sdtm_dataset_specializations_to_excel(workbook, "SDTM Dataset Specializations", df)
    workbook = write_dataframe_to_excel(workbook, "Domains", df_domain)
    workbook.save(args.excel_file)
    print(f"Excel file saved as {args.excel_file}")

    csv_file = args.excel_file.replace(".xlsx", ".csv")
    df.to_csv(csv_file, index=False, columns=HEADERS_SDTM)
    print(f"CSV file saved as {csv_file}")

if __name__ == "__main__":
    main()
