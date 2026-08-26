"""
Reads curation Excel workbooks by named range, replacing utilities/macros/readexcel.sas
and utilities/macros/get_excel_sheets.sas.

SAS's PROC IMPORT (dbms=excelcs) "range=" option accepts either a true workbook-scoped
named range (Excel "defined name"), or a bare/`$`-suffixed sheet reference (e.g. "BC_LB" or
"Subset Codelist Example$") meaning "the whole sheet by that name" when no defined name of
that name exists. In practice the curation workbooks in this repo use the latter almost
exclusively (they define no workbook-level named ranges at all) — read_named_range() tries
a defined name first, then falls back to treating the (`$`-stripped) name as a sheet name,
and returns a DataFrame whose columns match the header row, tagged with the same three
tracking columns readexcel.sas adds after PROC IMPORT: _tab_, _excel_file_, _record_.

Sheet-name matching is case-insensitive (e.g. "BC_SURROGATES$" resolves a sheet literally
named "BC_Surrogates"), matching PROC IMPORT's dbms=excelcs driver on Windows - confirmed by
a manifest job whose range and actual sheet name differ only in case
(utilities/manifests/bc/20260714_r18.yaml's BC_SURROGATES job against
curation/package18/R18_BC_Surrogates.xlsx's "BC_Surrogates" sheet).

Columns whose header cell is blank are dropped entirely (rather than kept as an empty-string
column name) - openpyxl's used-range detection can pick up trailing formatted-but-empty
columns well past the real data (confirmed against R18_BC_SDTM_SC.xlsx's "SDTM_SC" sheet,
which has ~400 such columns), and pandas can't concat/merge frames with duplicate ""
column labels. No converter or validator ever references a column by that name, so dropping
them is safe.
"""

import os
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def _normalize_header(value):
    """PROC IMPORT turns a header like "Subset Short Name" into the SAS variable name
    subset_short_name (spaces/punctuation -> underscore); the generate_yaml_from_* macros
    and get_subset_codelists.sas all reference columns by that lowercase snake_case form."""
    text = "" if value is None else str(value).strip()
    text = re.sub(r"[^0-9A-Za-z]+", "_", text).strip("_")
    return text.lower()


def _find_sheet_name_case_insensitive(sheet_names, bare_name):
    target = bare_name.lower()
    for sheet_name in sheet_names:
        if sheet_name.lower() == target:
            return sheet_name
    return None


def get_sheet_names(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_named_range(path, range_name):
    workbook = load_workbook(path, data_only=True)
    try:
        bare_name = range_name.rstrip("$")
        defined_name = workbook.defined_names.get(range_name) or workbook.defined_names.get(bare_name)

        if defined_name is not None:
            destinations = list(defined_name.destinations)
            if len(destinations) != 1:
                raise ValueError(
                    f"Named range '{range_name}' in {path} must resolve to exactly one "
                    f"sheet range, got {len(destinations)}"
                )
            sheet_title, cell_range = destinations[0]
            sheet = workbook[sheet_title]
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
            rows = list(
                sheet.iter_rows(
                    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
                )
            )
        else:
            sheet_name = _find_sheet_name_case_insensitive(workbook.sheetnames, bare_name)
            if sheet_name is None:
                raise KeyError(f"Named range/sheet '{range_name}' not found in {path}")
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise ValueError(f"Named range '{range_name}' in {path} is empty")

    header = [_normalize_header(cell) for cell in rows[0]]
    keep_positions = [i for i, name in enumerate(header) if name != ""]
    header = [header[i] for i in keep_positions]
    data_rows = [tuple(row[i] for i in keep_positions) for row in rows[1:]]
    df = pd.DataFrame(data_rows, columns=header)

    df["_tab_"] = range_name.replace("$", "")
    df["_excel_file_"] = os.path.basename(path)
    df["_record_"] = range(1, len(df) + 1)

    return df
